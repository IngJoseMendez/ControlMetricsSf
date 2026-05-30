# -*- coding: utf-8 -*-
"""
Utilidades estadísticas para el Sistema SPC - PMD y Cía S.C.A.
"""
import numpy as np
from scipy import stats
from config import SPC_CONSTANTS


def parse_value(text):
    """Convierte texto a float aceptando coma o punto decimal."""
    if not text or not text.strip():
        return None
    try:
        return float(text.strip().replace(',', '.'))
    except ValueError:
        return None


def compute_xbar_r(data):
    """
    Calcula medias y rangos de subgrupos.
    data: lista de listas de floats
    Retorna: (xbars, ranges)
    """
    xbars = [np.mean(row) for row in data]
    ranges = [max(row) - min(row) for row in data]
    return xbars, ranges


def compute_control_limits_xr(xbars, ranges, n):
    """
    Calcula límites de control X-barra y R.
    Retorna dict con llaves: xbar_bar, r_bar, sigma_est,
    ucl_x, cl_x, lcl_x, ucl_r, cl_r, lcl_r
    """
    c = SPC_CONSTANTS[n]
    xbar_bar = np.mean(xbars)
    r_bar    = np.mean(ranges)
    sigma_est = r_bar / c['d2']

    ucl_x = xbar_bar + 3 * sigma_est / np.sqrt(n)
    lcl_x = xbar_bar - 3 * sigma_est / np.sqrt(n)
    cl_x  = xbar_bar

    ucl_r = r_bar * c['D4']
    lcl_r = r_bar * c['D3']
    cl_r  = r_bar

    return {
        'xbar_bar':  xbar_bar,
        'r_bar':     r_bar,
        'sigma_est': sigma_est,
        'ucl_x': ucl_x, 'cl_x': cl_x, 'lcl_x': lcl_x,
        'ucl_r': ucl_r, 'cl_r': cl_r, 'lcl_r': lcl_r,
        'd2': c['d2'], 'D3': c['D3'], 'D4': c['D4'], 'A2': c['A2'],
    }


def detect_out_of_control(values, ucl, lcl):
    """Retorna lista de índices fuera de control (base-0)."""
    return [i for i, v in enumerate(values) if v > ucl or v < lcl]


def compute_control_limits_c(defects):
    """
    Calcula límites de control para carta C.
    Retorna dict con c_bar, ucl, lcl.
    """
    c_bar = np.mean(defects)
    ucl = c_bar + 3 * np.sqrt(c_bar)
    lcl = max(0.0, c_bar - 3 * np.sqrt(c_bar))
    return {'c_bar': c_bar, 'ucl': ucl, 'lcl': lcl}


def normality_tests(data):
    """
    Ejecuta pruebas de normalidad.
    Retorna dict con resultados de AD, SW y KS.
    """
    data = np.array(data)
    results = {}

    # Anderson-Darling
    try:
        # scipy >= 1.17 soporta method='interpolate' con pvalue directo
        ad_result = stats.anderson(data, dist='norm', method='interpolate')
        ad_stat  = ad_result.statistic
        ad_pval  = float(ad_result.pvalue)
        ad_cv_5  = None
        ad_normal = ad_pval > 0.05
    except TypeError:
        # scipy < 1.17 (API antigua)
        ad_result = stats.anderson(data, dist='norm')
        ad_stat   = ad_result.statistic
        ad_cv_5   = ad_result.critical_values[2]
        ad_pval   = _ad_pvalue(ad_stat, len(data))
        ad_normal = ad_stat < ad_cv_5
    results['AD'] = {
        'statistic': ad_stat,
        'p_value':   ad_pval,
        'cv_5pct':   ad_cv_5,
        'normal':    ad_normal,
    }

    # Shapiro-Wilk (más fiable para n < 50)
    sw_stat, sw_pval = stats.shapiro(data)
    results['SW'] = {
        'statistic': sw_stat,
        'p_value':   sw_pval,
        'normal':    sw_pval > 0.05,
    }

    # Kolmogorov-Smirnov (normalizado)
    ks_stat, ks_pval = stats.kstest(data, 'norm', args=(data.mean(), data.std(ddof=1)))
    results['KS'] = {
        'statistic': ks_stat,
        'p_value':   ks_pval,
        'normal':    ks_pval > 0.05,
    }

    return results


def _ad_pvalue(ad_stat, n):
    """Aproximación del p-valor para la prueba Anderson-Darling."""
    # Ajuste por tamaño de muestra
    z = ad_stat * (1 + 0.75 / n + 2.25 / n**2)
    if z >= 0.6:
        p = np.exp(1.2937 - 5.709 * z + 0.0186 * z**2)
    elif z >= 0.34:
        p = np.exp(0.9177 - 4.279 * z - 1.38 * z**2)
    elif z >= 0.2:
        p = 1 - np.exp(-8.318 + 42.796 * z - 59.938 * z**2)
    else:
        p = 1 - np.exp(-13.436 + 101.14 * z - 223.73 * z**2)
    return float(np.clip(p, 0.0, 1.0))


def compute_capability(data, lie, lse, sigma=None):
    """
    Calcula índices de capacidad Cp, Cpk, Cpu, Cpl y %PNC.
    Si sigma es None se calcula desde los datos.
    """
    mu = np.mean(data)
    if sigma is None:
        sigma = np.std(data, ddof=1)

    cp  = (lse - lie) / (6 * sigma)
    cpu = (lse - mu)  / (3 * sigma)
    cpl = (mu  - lie) / (3 * sigma)
    cpk = min(cpu, cpl)

    # % no conformes (usando distribución normal)
    z_sup = (lse - mu) / sigma
    z_inf = (lie - mu) / sigma
    pnc_sup  = 1 - stats.norm.cdf(z_sup)
    pnc_inf  = stats.norm.cdf(z_inf)
    pnc_total = pnc_sup + pnc_inf

    # Nivel sigma
    sigma_level = abs(min(z_sup, -z_inf))

    return {
        'mu': mu, 'sigma': sigma,
        'cp': cp, 'cpu': cpu, 'cpl': cpl, 'cpk': cpk,
        'pnc_sup': pnc_sup, 'pnc_inf': pnc_inf, 'pnc_total': pnc_total,
        'sigma_level': sigma_level,
        'z_sup': z_sup, 'z_inf': z_inf,
    }


def compute_oc_curve(n, c, p_range=None):
    """
    Calcula la Curva de Operación Característica (OC) para un plan (n, c).
    Usa distribución binomial.
    Retorna (p_values, pa_values).
    """
    if p_range is None:
        p_range = np.linspace(0, 0.30, 300)
    pa = stats.binom.cdf(c, n, p_range)
    return p_range, pa


def compute_arl(ucl, lcl, xbar_bar, sigma, n, delta=0):
    """
    Calcula ARL y ATS para un corrimiento delta (en unidades de sigma).
    delta: corrimiento de la media en libras.
    Retorna dict con beta (P(no detectar)), arl, ats.
    """
    mu_nuevo = xbar_bar + delta
    se = sigma / np.sqrt(n)
    beta = stats.norm.cdf((ucl - mu_nuevo) / se) - stats.norm.cdf((lcl - mu_nuevo) / se)
    beta = float(np.clip(beta, 0.0, 1.0))
    poder = 1 - beta
    arl = 1 / poder if poder > 0 else float('inf')
    return {'beta': beta, 'poder': poder, 'arl': arl}


# ── Reglas de Western Electric / Nelson ──────────────────────────────────────

def western_electric_rules(values, ucl, cl, lcl):
    """
    Detecta violaciones de las 6 reglas más usadas en SPC (Nelson/WE).
    Retorna dict {1: [índices], ..., 6: [índices]}.
    """
    v = np.asarray(values, dtype=float)
    n = len(v)
    sigma = (ucl - cl) / 3.0
    viol = {i: [] for i in range(1, 7)}

    for i in range(n):
        if v[i] > ucl or v[i] < lcl:
            viol[1].append(i)

    for i in range(2, n):
        w = v[i-2:i+1]
        if sum(x > cl + 2*sigma for x in w) >= 2 or \
           sum(x < cl - 2*sigma for x in w) >= 2:
            viol[2].append(i)

    for i in range(4, n):
        w = v[i-4:i+1]
        if sum(x > cl + sigma for x in w) >= 4 or \
           sum(x < cl - sigma for x in w) >= 4:
            viol[3].append(i)

    for i in range(7, n):
        w = v[i-7:i+1]
        if all(x > cl for x in w) or all(x < cl for x in w):
            viol[4].append(i)

    for i in range(5, n):
        w = v[i-5:i+1]
        if all(w[j] < w[j+1] for j in range(5)) or \
           all(w[j] > w[j+1] for j in range(5)):
            viol[5].append(i)

    for i in range(14, n):
        w = v[i-14:i+1]
        if all(cl - sigma < x < cl + sigma for x in w):
            viol[6].append(i)

    return {k: sorted(set(lst)) for k, lst in viol.items()}


# ── Carta I-MR ───────────────────────────────────────────────────────────────

def compute_imr(data):
    """Individual values and moving ranges (span=2)."""
    ind = list(data)
    mr  = [abs(ind[i] - ind[i-1]) for i in range(1, len(ind))]
    return ind, mr


def compute_control_limits_imr(individuals, moving_ranges):
    """Límites de control I-MR (Montgomery 2020, d2=1.128, D4=3.267)."""
    x_bar    = float(np.mean(individuals))
    mr_bar   = float(np.mean(moving_ranges))
    sigma_est = mr_bar / 1.128
    return {
        'x_bar': x_bar, 'mr_bar': mr_bar, 'sigma_est': sigma_est,
        'ucl_i': x_bar + 3*sigma_est, 'cl_i': x_bar, 'lcl_i': x_bar - 3*sigma_est,
        'ucl_mr': 3.267*mr_bar,       'cl_mr': mr_bar, 'lcl_mr': 0.0,
    }


# ── Carta X̄-S ────────────────────────────────────────────────────────────────

def compute_xbar_s(data):
    """Medias y desviaciones estándar de subgrupos."""
    return ([float(np.mean(r)) for r in data],
            [float(np.std(r, ddof=1)) for r in data])


def compute_control_limits_xs(xbars, stds, n):
    """Límites de control X̄-S usando constantes c4, A3, B3, B4."""
    c4   = SPC_CONSTANTS[n]['c4']
    root = np.sqrt(max(0.0, 1 - c4**2))
    A3   = 3.0 / (c4 * np.sqrt(n))
    B4   = 1.0 + 3.0*root/c4
    B3   = max(0.0, 1.0 - 3.0*root/c4)
    xb   = float(np.mean(xbars))
    sb   = float(np.mean(stds))
    return {
        'xbar_bar': xb, 's_bar': sb, 'sigma_est': sb/c4,
        'ucl_x': xb + A3*sb, 'cl_x': xb, 'lcl_x': xb - A3*sb,
        'ucl_s': B4*sb,       'cl_s': sb, 'lcl_s': B3*sb,
        'A3': A3, 'B3': B3, 'B4': B4, 'c4': c4,
    }


# ── Cartas por atributos (p, np, u) ──────────────────────────────────────────

def compute_p_chart(defective_counts, sample_sizes):
    """Carta p – proporción no conforme (n puede variar)."""
    np_arr = np.array(defective_counts, dtype=float)
    n_arr  = np.array(sample_sizes,     dtype=float)
    p_bar  = float(np_arr.sum() / n_arr.sum())
    p_vals = (np_arr / n_arr).tolist()
    ucl = (p_bar + 3*np.sqrt(p_bar*(1-p_bar)/n_arr)).tolist()
    lcl = np.maximum(0, p_bar - 3*np.sqrt(p_bar*(1-p_bar)/n_arr)).tolist()
    return {'p_bar': p_bar, 'p_values': p_vals,
            'ucl_p': ucl, 'cl_p': p_bar, 'lcl_p': lcl}


def compute_np_chart(defective_counts, n):
    """Carta np – número no conforme (n constante)."""
    arr   = np.array(defective_counts, dtype=float)
    np_bar = float(arr.mean())
    p_bar  = np_bar / n
    std    = np.sqrt(np_bar*(1 - p_bar))
    return {'np_bar': np_bar, 'p_bar': p_bar, 'np_values': arr.tolist(),
            'ucl_np': np_bar+3*std, 'cl_np': np_bar,
            'lcl_np': max(0.0, np_bar-3*std)}


def compute_u_chart(defect_counts, sample_sizes):
    """Carta u – defectos por unidad (n puede variar)."""
    c_arr = np.array(defect_counts,  dtype=float)
    n_arr = np.array(sample_sizes,   dtype=float)
    u_bar  = float(c_arr.sum() / n_arr.sum())
    u_vals = (c_arr / n_arr).tolist()
    ucl = (u_bar + 3*np.sqrt(u_bar/n_arr)).tolist()
    lcl = np.maximum(0, u_bar - 3*np.sqrt(u_bar/n_arr)).tolist()
    return {'u_bar': u_bar, 'u_values': u_vals,
            'ucl_u': ucl, 'cl_u': u_bar, 'lcl_u': lcl}


# ── Tests de independencia ────────────────────────────────────────────────────

def durbin_watson(residuals):
    """Estadístico Durbin-Watson (2 ≈ sin autocorrelación)."""
    r  = np.asarray(residuals, dtype=float)
    dw = float(np.sum(np.diff(r)**2) / np.sum((r - r.mean())**2))
    if dw < 1.5:
        interp, indep = 'Autocorrelación positiva', False
    elif dw > 2.5:
        interp, indep = 'Autocorrelación negativa', False
    else:
        interp, indep = 'Sin autocorrelación significativa', True
    return {'statistic': dw, 'interpretation': interp, 'independent': indep}


def runs_test(data):
    """Prueba de rachas sobre la mediana."""
    d      = np.asarray(data, dtype=float)
    median = float(np.median(d))
    signs  = d > median
    n1, n2 = int(signs.sum()), int((~signs).sum())
    runs   = 1 + int(np.sum(signs[:-1] != signs[1:]))
    if n1 == 0 or n2 == 0 or (n1+n2) <= 2:
        return {'runs': runs, 'z': 0.0, 'p_value': 1.0, 'independent': True,
                'n_above': n1, 'n_below': n2}
    mu_r  = 2*n1*n2/(n1+n2) + 1
    var_r = (mu_r-1)*(mu_r-2) / max(1, n1+n2-1)
    z     = (runs - mu_r) / np.sqrt(var_r)
    p     = float(2*(1 - stats.norm.cdf(abs(z))))
    return {'runs': runs, 'n_above': n1, 'n_below': n2,
            'z': float(z), 'p_value': p, 'independent': p > 0.05}


def acf_lag1(data):
    """Autocorrelación lag-1 con test aproximado."""
    d    = np.asarray(data, dtype=float)
    mean = d.mean()
    num  = float(np.sum((d[:-1]-mean)*(d[1:]-mean)))
    den  = float(np.sum((d-mean)**2))
    r1   = num/den if den > 0 else 0.0
    n    = len(d)
    z    = r1 * np.sqrt(n)
    p    = float(2*(1 - stats.norm.cdf(abs(z))))
    return {'r1': r1, 'z': z, 'p_value': p, 'independent': p > 0.05}


def compute_power_curve(ucl, lcl, xbar_bar, sigma, n, delta_sigma_range):
    """
    Curva de potencia para carta X-barra.
    delta_sigma_range: iterable de corrimientos en unidades de sigma.
    Retorna lista de dicts: delta_sigma, delta_real, poder, beta, arl.
    """
    results = []
    for d in np.asarray(delta_sigma_range):
        r = compute_arl(ucl, lcl, xbar_bar, sigma, n, delta=float(d) * sigma)
        results.append({
            'delta_sigma': float(d),
            'delta_real':  float(d) * sigma,
            'poder': r['poder'],
            'beta':  r['beta'],
            'arl':   r['arl'],
        })
    return results


def summary_stats(data):
    """Estadísticas descriptivas básicas."""
    data = np.array(data)
    return {
        'n':      len(data),
        'mean':   float(np.mean(data)),
        'std':    float(np.std(data, ddof=1)),
        'median': float(np.median(data)),
        'min':    float(np.min(data)),
        'max':    float(np.max(data)),
        'q1':     float(np.percentile(data, 25)),
        'q3':     float(np.percentile(data, 75)),
        'cv':     float(np.std(data, ddof=1) / np.mean(data) * 100),
    }


# ─── Lectura desde Excel ─────────────────────────────────────────────────────

def excel_read_single_col(filepath, sheet_idx=0, skip_header=True):
    """
    Lee una columna (col A) de valores numéricos desde Excel.
    Formato esperado: fila 1 = encabezado (opcional), filas siguientes = valores.
    Retorna lista de floats.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_idx]
    values = []
    start = 2 if skip_header else 1
    for row in ws.iter_rows(min_row=start, max_col=1, values_only=True):
        cell = row[0]
        if cell is None:
            continue
        val = parse_value(str(cell))
        if val is not None:
            values.append(val)
    wb.close()
    return values


def excel_read_subgroups(filepath, sheet_idx=0, skip_header=True):
    """
    Lee subgrupos desde Excel. Cada fila = un subgrupo, columnas = observaciones.
    Formato: fila 1 = encabezados (Subgrupo, X1, X2, ...), filas 2+ = datos.
    Si la primera columna contiene el número de subgrupo se omite automáticamente.
    Retorna lista de listas de floats.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_idx]
    data = []
    start = 2 if skip_header else 1
    for row in ws.iter_rows(min_row=start, values_only=True):
        row_vals = []
        for cell in row:
            if cell is None:
                break
            val = parse_value(str(cell))
            if val is not None:
                row_vals.append(val)
        if not row_vals:
            break
        data.append(row_vals)
    wb.close()
    return data


def excel_read_defects(filepath):
    """
    Lee datos para Carta C y Pareto.
    Hoja 1 "CartaC": col A = defectos por caja (con encabezado).
    Hoja 2 "Pareto": col A = tipo de defecto, col B = frecuencia (con encabezado).
    Retorna dict: {'cajas': [...], 'tipos': {name: freq}}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result = {'cajas': [], 'tipos': {}}

    ws1 = wb.worksheets[0]
    for row in ws1.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is None:
            break
        val = parse_value(str(row[0]))
        if val is not None:
            result['cajas'].append(int(round(val)))

    if len(wb.worksheets) >= 2:
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] is None:
                break
            tipo = str(row[0]).strip()
            frec = parse_value(str(row[1])) if row[1] is not None else None
            if tipo and frec is not None:
                result['tipos'][tipo] = int(round(frec))

    wb.close()
    return result


def excel_read_muestreo(filepath):
    """
    Lee parámetros del plan de muestreo.
    Formato: col A = nombre del parámetro, col B = valor.
    Claves reconocidas: n_lote, pallets, cajas_pallet, nac.
    Retorna dict de parámetros.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    params = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is None:
            continue
        key = str(row[0]).strip().lower().replace(' ', '_').replace('-', '_')
        val = row[1]
        if val is not None:
            params[key] = val
    wb.close()
    return params


def excel_create_template(filepath, template_type):
    """
    Crea una plantilla Excel vacía con el formato correcto para cada módulo.
    template_type: 'single_col' | 'subgroups' | 'defects' | 'muestreo'
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    header_fill = PatternFill('solid', fgColor='1B5E20')
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    def style_header(cell):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    if template_type == 'single_col':
        ws = wb.active
        ws.title = 'Datos'
        ws['A1'] = 'Peso (lb)'
        style_header(ws['A1'])
        ws.column_dimensions['A'].width = 14
        for i, ex in enumerate([3.22, 3.18, 3.25, 3.30, 3.15], start=2):
            ws.cell(row=i, column=1, value=ex)

    elif template_type == 'subgroups':
        ws = wb.active
        ws.title = 'Subgrupos'
        headers = ['Subgrupo', 'X1', 'X2', 'X3', 'X4', 'X5', 'X6']
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=h)
            style_header(c)
            ws.column_dimensions[c.column_letter].width = 10
        example = [1, 3.22, 3.18, 3.25, 3.30, 3.15, 3.28]
        for j, v in enumerate(example, start=1):
            ws.cell(row=2, column=j, value=v)

    elif template_type == 'defects':
        ws1 = wb.active
        ws1.title = 'CartaC'
        ws1['A1'] = 'Defectos por caja'
        style_header(ws1['A1'])
        ws1.column_dimensions['A'].width = 20
        for i, v in enumerate([2, 1, 0, 1, 2], start=2):
            ws1.cell(row=i, column=1, value=v)

        ws2 = wb.create_sheet('Pareto')
        for j, h in enumerate(['Tipo de Defecto', 'Frecuencia'], start=1):
            c = ws2.cell(row=1, column=j, value=h)
            style_header(c)
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 12
        examples = [('Speckling', 5), ('Trips', 4), ('Mancha de madurez', 3)]
        for i, (t, f) in enumerate(examples, start=2):
            ws2.cell(row=i, column=1, value=t)
            ws2.cell(row=i, column=2, value=f)

    elif template_type == 'muestreo':
        ws = wb.active
        ws.title = 'Parametros'
        for j, h in enumerate(['Parámetro', 'Valor'], start=1):
            c = ws.cell(row=1, column=j, value=h)
            style_header(c)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        params = [('n_lote', 960), ('pallets', 20), ('cajas_pallet', 48), ('nac', 4.0)]
        for i, (k, v) in enumerate(params, start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

    wb.save(filepath)


def milstd_code_letter(lot_size):
    """Determina la letra código MIL-STD-105E (Nivel de Inspección II)."""
    from config import MILSTD_CODE_LETTER
    for (lo, hi), letter in MILSTD_CODE_LETTER.items():
        if lo <= lot_size <= hi:
            return letter
    return 'L'


def milstd_plan(lot_size):
    """Retorna el plan (n, Ac, Re) para AQL=4%, Inspección Normal."""
    from config import MILSTD_PLANS
    letter = milstd_code_letter(lot_size)
    plan = MILSTD_PLANS.get(letter, (80, 7, 8))
    return letter, plan
